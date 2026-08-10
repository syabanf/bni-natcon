#!/usr/bin/env python3
"""Regenerates the QA scenario markdown from the workbook.

The workbook is the source of truth; this keeps the readable copy in step so
the two can never drift apart:

    python3 scripts/qa_md_from_xlsx.py
"""
import openpyxl

SRC = "docs/qa/natcon2026-qa-scenarios.xlsx"
OUT = "docs/qa/qa-scenarios.md"

wb = openpyxl.load_workbook(SRC)
lines = [
    "# BNI Natcon 2026 — QA Scenario Test Pack",
    "",
    "> Generated from [`natcon2026-qa-scenarios.xlsx`](natcon2026-qa-scenarios.xlsx) —",
    "> edit the workbook, then run `python3 scripts/qa_md_from_xlsx.py`.",
    "",
]

# Read-me sheet, verbatim.
for row in wb["00 Read me"].iter_rows(min_row=2, max_col=1, values_only=True):
    text = (row[0] or "").strip()
    if not text:
        lines.append("")
    elif text in ("How to use", "Environment", "Before you start", "Sheets") or text.startswith("Accounts"):
        lines.append(f"## {text}")
    else:
        lines.append(text)
lines.append("")

counts = []
for name in wb.sheetnames:
    if name.startswith(("00", "08")):
        continue
    ws = wb[name]
    cases = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0]]
    if not cases:
        continue
    counts.append((name, len(cases), sum(1 for c in cases if c[1] == "P1")))
    lines += [f"## {name[3:]}", "", f"*{ws.cell(1, 1).value}*", "",
              "| ID | Pri | Precondition | Steps | Expected result |",
              "|---|---|---|---|---|"]
    for c in cases:
        cells = [str(x or "").replace("|", "\\|").replace("\n", " ") for x in c[:5]]
        lines.append("| " + " | ".join(cells) + " |")
    lines.append("")

lines += ["## Test data", "", "| What | Value | Notes |", "|---|---|---|"]
for row in wb["08 Test data"].iter_rows(min_row=2, values_only=True):
    if row[0]:
        lines.append("| " + " | ".join(str(x or "").replace("|", "\\|") for x in row[:3]) + " |")
lines.append("")

total = sum(n for _, n, _ in counts)
p1 = sum(p for _, _, p in counts)
lines += ["## Coverage", "", "| Section | Cases | P1 |", "|---|---|---|"]
for name, n, p in counts:
    lines.append(f"| {name[3:]} | {n} | {p} |")
lines += [f"| **Total** | **{total}** | **{p1}** |", ""]

open(OUT, "w").write("\n".join(lines))
print(f"wrote {OUT} — {total} cases, {p1} P1")

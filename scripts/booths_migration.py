#!/usr/bin/env python3
"""Turn the committee's booth sheet into the booth migration.

The booth list is event master data, so it lives in a migration: a laptop, a
staging box and production all end up with the same booths without anyone
remembering to import a spreadsheet.

    pip install openpyxl
    python3 scripts/booths_migration.py "~/Downloads/Data Peserta & Booth ... .xlsx"

Rewrites backend/internal/repository/postgres/migrations/0037_booths.sql.
Re-run it whenever the sheet changes, then restart the API.

Expected columns on the "Booth" sheet: Booth Number · Company Name ·
Business Classification · Name · BNI Chapter. That is the committee's own
export, unedited.

WHAT SURVIVES A RE-RUN
----------------------
An exhibitor is matched on its COMPANY NAME, not its stand, because the floor
plan has been renumbered twice already. So when the committee moves a company
to another stand, this migration MOVES it — the exhibitor keeps its identity,
its scanner login and every scan it has collected — instead of deleting one
booth and creating another. An exhibitor that leaves the sheet is removed only
when nobody has scanned it: a database in use is never quietly emptied.

Two booth numbers on one row ("A18 & 20", "A47-48") means one exhibitor
holding two positions on the floor plan. That is ONE exhibitor: one entry in
the passport, one stamp, one scanner login, and one QR printed twice — once
for each sign. Splitting them would let an attendee collect two stamps from
the same company and count it twice towards the draw's booth minimum.

LOGOS
-----
The sheet carries no artwork, so logos come from scripts/booth-logos.json —
company name to file under frontend/public/logos/. That mapping outlives the
floor plan: the files are named after the company, never the stand. A company
the pack has no logo for keeps its initials, which is what the UI falls back
to.
"""

import argparse
import json
import pathlib
import re
import sys

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "backend/internal/repository/postgres/migrations/0037_booths.sql"
LOGOS = ROOT / "scripts/booth-logos.json"

PLACEHOLDER_HASH = "$2a$10$SEEDPLACEHOLDERSEEDPLACEHOLDERSEEDPLACEHOLDERSEEDPLACEH"


# Typos in the committee's own sheet that we correct on the way through.
#
# The sheet is the source of truth for WHO is on the floor and WHERE, and this
# is deliberately not a place to restyle names — only to fix a company's name
# being wrong. Keyed on what the sheet says, so re-running after the committee
# fixes their own copy simply stops matching.
CORRECTIONS = {
    # Missing the P of PT, and it is a stand shared by two companies.
    "T Royal Medicalink Pharmalab & PT Aroma Bathi Indonesia":
        "PT Royal Medicalink Pharmalab & PT Aroma Bathi Indonesia",
}


def initials(name: str) -> str:
    """Mirror of tenantInitials in the Go usecase, so a booth created here and
    one created through the admin panel look the same."""
    out = "".join(w[0] for w in name.split())
    return out[:2].upper()


def key(name: str) -> str:
    """How an exhibitor is recognised across sheets: letters and digits only.
    'WIT.id' and 'WIT.ID' are the same company; so are 'Alpha leaders' and
    'ALPHA LEADERS'. Mirrored in SQL as regexp_replace(name,'[^a-zA-Z0-9]','')."""
    return re.sub(r"[^a-z0-9]", "", name.lower())


def login_email(booth: str) -> str:
    """Mirror of domain.TenantLoginEmail — the address on the briefing sheet."""
    return f"booth-{re.sub(r'[^a-z0-9]', '', booth.lower())}@natcon.id"


def q(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def booth_codes(cell: str):
    """'A18 & 20' -> ['A18', 'A20'];  'A47-48' -> ['A47', 'A48'].

    The second half is often written as a bare number, so it borrows the
    letter from the first.
    """
    parts = [p.strip() for p in re.split(r"\s*(?:&|,|\band\b|-|–)\s*", cell) if p.strip()]
    if not parts:
        return []
    prefix = re.match(r"[A-Za-z]+", parts[0])
    prefix = prefix.group(0).upper() if prefix else ""
    out = []
    for p in parts:
        p = p.upper()
        out.append(p if re.match(r"[A-Z]", p) else prefix + p)
    return out


def read_booths(path: pathlib.Path, sheet_name: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(header)}
    for needed in ("Booth Number", "Company Name"):
        if needed not in col:
            sys.exit(f"column {needed!r} missing from sheet {ws.title!r} — got {header}")

    logos = json.loads(LOGOS.read_text(encoding="utf-8")) if LOGOS.exists() else {}
    by_key = {key(company): url for company, url in logos.items()}

    booths, seen = [], set()
    for r in rows[1:]:
        def cell(name):
            i = col.get(name)
            return str(r[i]).strip() if i is not None and r[i] is not None else ""

        number, company = cell("Booth Number"), cell("Company Name")
        company = CORRECTIONS.get(company, company)
        if not number:
            # The sheet ends with a row naming the organiser and no stand.
            continue
        if not company:
            # A stand with no company names only its exhibitor (A49, rev. 31
            # Aug): the person IS the exhibitor, and skipping the row would
            # drop a real booth from the passport.
            company = cell("Name")
            if not company:
                continue
        codes = booth_codes(number)
        if not codes or codes[0] in seen:
            continue
        seen.add(codes[0])
        # Stand A is the exhibition floor; B and C are the sponsor stands, and
        # the sponsors lead the passport ahead of the floor.
        kind = "booth" if codes[0][0].upper() == "A" else "sponsor"
        booths.append({
            "key": key(company),
            # Two stands, one exhibitor: the label names both so the passport
            # and the printed sign say what the floor plan says, while the
            # login and the QR follow the first code — there is only one.
            "booth": " & ".join(codes),
            "company": company,
            "category": cell("Business Classification"),
            "initials": initials(company),
            "kind": kind,
            "email": login_email(codes[0]),
            "logo": by_key.get(key(company), ""),
            "contact": cell("Name"),
            "chapter": cell("BNI Chapter"),
        })
    return booths


FIELDS = ("key", "booth", "company", "category", "initials", "kind", "email",
          "logo", "contact", "chapter")


def values_block(booths, indent="    ") -> str:
    return ",\n".join(indent + "(" + ", ".join(q(b[f]) for f in FIELDS) + ")"
                      for b in booths)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("sheet")
    ap.add_argument("--sheet-name", default="Booth",
                    help="worksheet holding the booth list (default: Booth)")
    args = ap.parse_args()

    src = pathlib.Path(args.sheet).expanduser()
    booths = read_booths(src, args.sheet_name)
    if not booths:
        sys.exit("no booth rows found")

    block = values_block(booths)
    codes = ", ".join(q(b["booth"]) for b in booths)
    sponsors = sum(1 for b in booths if b["kind"] == "sponsor")
    logos = sum(1 for b in booths if b["logo"])
    cols = "(key, booth, company, category, initials, kind, email, logo, contact, chapter)"

    sql = f"""-- The exhibitor floor: {len(booths) - sponsors} booths and {sponsors} sponsors, from the
-- committee's booth sheet.
-- GENERATED by scripts/booths_migration.py — edit the sheet, not this file.
--
-- This replaces every earlier booth migration. Those were built from sheets
-- the committee has since redrawn, and re-running them would put exhibitors
-- back on stands they have left.
--
-- An exhibitor is recognised by its COMPANY NAME, not its stand: the floor
-- plan has been renumbered twice, and a company that moves must keep its
-- identity, its scanner login and every scan it has collected. So a move is a
-- MOVE here, never a delete and a re-create.
--
-- Scanner logins follow the stand (booth-<code>@natcon.id) and the first
-- password follows the company and the stand together, so both change when an
-- exhibitor moves. The rows below carry a placeholder hash nobody can sign in
-- with; the API's seeder rewrites it, per booth, on the next start.

-- ---------------------------------------------------------------- moves
-- Park every exhibitor whose stand changed under a temporary address first:
-- the logins are unique, and A20 -> A22 while A22 -> A23 would collide
-- halfway through a single statement.
UPDATE users u
SET email = 'moving-' || u.id || '@natcon.id'
FROM tenants t, (VALUES
{block}
) AS v {cols}
WHERE t.owner_user_id = u.id
  AND lower(regexp_replace(t.name, '[^a-zA-Z0-9]', '', 'g')) = v.key
  AND t.booth <> v.booth;

UPDATE tenants t
SET booth = v.booth
FROM (VALUES
{block}
) AS v {cols}
WHERE lower(regexp_replace(t.name, '[^a-zA-Z0-9]', '', 'g')) = v.key
  AND t.booth <> v.booth;

UPDATE users u
SET email = v.email
FROM tenants t, (VALUES
{block}
) AS v {cols}
WHERE t.owner_user_id = u.id
  AND u.email LIKE 'moving-%'
  AND lower(regexp_replace(t.name, '[^a-zA-Z0-9]', '', 'g')) = v.key;

-- ---------------------------------------------------------------- arrivals
-- Scanner logins first: one tenant user per booth.
INSERT INTO users (name, email, password_hash, role, company)
SELECT v.company, v.email, '{PLACEHOLDER_HASH}', 'tenant', v.company
FROM (VALUES
{block}
) AS v {cols}
WHERE NOT EXISTS (SELECT 1 FROM users u WHERE u.email = v.email)
  AND NOT EXISTS (SELECT 1 FROM tenants t WHERE t.booth = v.booth);

INSERT INTO tenants (name, category, booth, initials, kind, description,
                     contact_name, chapter, logo_url, owner_user_id)
SELECT v.company, v.category, v.booth, v.initials, v.kind, '',
       v.contact, v.chapter, v.logo, u.id
FROM (VALUES
{block}
) AS v {cols}
JOIN users u ON u.email = v.email
WHERE NOT EXISTS (SELECT 1 FROM tenants t WHERE t.booth = v.booth);

-- ---------------------------------------------------------------- details
-- An exhibitor already on the floor plan follows the sheet: the committee
-- edits there, not here. The logo is the one exception — it comes from the
-- artwork pack, so an empty entry never wipes a logo already in place.
UPDATE tenants t
SET name = v.company, category = v.category, initials = v.initials,
    kind = v.kind, contact_name = v.contact, chapter = v.chapter,
    logo_url = CASE WHEN v.logo <> '' THEN v.logo ELSE t.logo_url END
FROM (VALUES
{block}
) AS v {cols}
WHERE t.booth = v.booth;

-- A crew that has not signed in yet goes back to the placeholder, so the
-- seeder derives their first password from the stand they are on NOW. A crew
-- that already chose its own password is left alone.
UPDATE users u
SET password_hash = '{PLACEHOLDER_HASH}'
FROM tenants t
WHERE t.owner_user_id = u.id
  AND u.role = 'tenant'
  AND u.must_set_password = true;

-- ---------------------------------------------------------------- departures
-- Anything else that calls itself a booth or a sponsor is not on this floor
-- plan: a leftover from the sheet this migration replaces. The tenant row goes
-- first; its scanner login is behind a foreign key.
WITH removed AS (
    DELETE FROM tenants
    WHERE booth NOT IN ({codes})
      AND NOT EXISTS (SELECT 1 FROM visits v WHERE v.tenant_id = tenants.id)
    RETURNING owner_user_id
)
DELETE FROM users u
USING removed r
WHERE u.id = r.owner_user_id AND u.role = 'tenant';
"""
    OUT.write_text(sql, encoding="utf-8")
    print(f"{OUT.relative_to(ROOT)} — {len(booths)} exhibitors "
          f"({sponsors} sponsors), {logos} logos, from {src.name}")
    for b in booths:
        if not b["logo"]:
            print(f"  no logo: {b['booth']:<9} {b['company']}")


if __name__ == "__main__":
    main()

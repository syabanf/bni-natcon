#!/usr/bin/env python3
"""Turn the ticketing export into the attendee seed migration.

    pip install openpyxl
    python3 scripts/attendees_migration.py "~/Downloads/Data Peserta & Booth ... .xlsx"

Writes backend/internal/repository/postgres/migrations/0038_attendees.sql.

WHAT IT WRITES
--------------
· every chapter the sheet mentions, so the master list matches the sheet;
· one member per TICKET — the ticket number is the identity, and it is what
  the attendee's QR carries. A buyer holding several tickets becomes several
  attendees, which is the point: a booth owner buying four crew passes is
  four people through the door, and the sheet never collected the crew's
  names;
· the committee's single first password for everyone, written as the
  seeder's placeholder rather than hashed here — the API turns it into
  SEED_PASSWORD on the next start. must_set_password stays true, so the app
  still makes each attendee choose their own on first sign-in.

THE COMPANY COLUMN
------------------
The committee's newer export dropped Company Name and put Business
Classification and Industry in its place. A company already known for a
ticket is carried across from scripts/attendee-companies.json rather than
thrown away; someone registering for the first time simply has no company on
file, and the field stays empty. Nothing is guessed — a classification is not
an employer, and inventing one would put a wrong name on a booth's lead sheet.

Re-running is safe: a ticket already in the database is left exactly as it is,
keeping its password, its booth stamps and its class registrations. A ticket
that has LEFT the sheet is deleted, but only when that attendee has no scans,
no class seat and no check-in against their name — a database in use is never
quietly emptied.
"""

import argparse
import json
import pathlib
import re
import sys
import unicodedata

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent

# The hash nobody can sign in with. The API's seeder rewrites it with
# SEED_PASSWORD on the next start — see SeedIfEmpty.
PLACEHOLDER_HASH = "$2a$10$SEEDPLACEHOLDERSEEDPLACEHOLDERSEEDPLACEHOLDERSEEDPLACEH"
OUT = ROOT / "backend/internal/repository/postgres/migrations/0038_attendees.sql"
COMPANIES = ROOT / "scripts/attendee-companies.json"

COLUMNS = {
    "ticket": "Ticket Number",
    "first": "First Name",
    "last": "Last Name",
    "email": "Email",
    "phone": "Phone",
    "chapter": "Bni Chapter",
    "classification": "Business Classification",
}


# Chapters the committee has told us are wrong in their own export.
#
# Keyed on the TICKET NUMBER, not the name: a ticket is one person, and two
# attendees can share a name. Correcting here rather than in the database
# means a re-import keeps the fix; and the day the committee fixes their own
# sheet, the entry simply stops matching anything.
#
# Every line needs somebody to have said so — this is not a place to guess a
# chapter from a name.
CHAPTER_CORRECTIONS = {
    # Stephanie Safitri Jusuf — sheet says Amplify, committee says Prestige.
    "16798-2556D8630": "Prestige",
}


def q(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def name_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", unicodedata.normalize("NFKD", name or "").lower())


def normalize_phone(raw: str) -> str:
    """Mirror of normalizePhone in admin/src/excel.js — the sheet writes
    numbers as text with Excel's leading apostrophe."""
    p = re.sub(r"[\s.\-()]", "", raw.strip().lstrip("'"))
    if not p:
        return ""
    if p.startswith("+"):
        return "+" + re.sub(r"\D", "", p[1:])
    p = re.sub(r"\D", "", p)
    if p.startswith("0"):
        return "+62" + p[1:]
    if p.startswith("62"):
        return "+" + p
    return p


def read_rows(path: pathlib.Path, sheet_name: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    ix = {}
    for key, column in COLUMNS.items():
        if column not in header:
            sys.exit(f"column {column!r} missing from sheet {ws.title!r} — got {header}")
        ix[key] = header.index(column)

    known = json.loads(COMPANIES.read_text(encoding="utf-8")) if COMPANIES.exists() else {}
    by_ticket, by_name = known.get("by_ticket", {}), known.get("by_name", {})

    out, seen, carried = [], set(), [0, 0]
    for r in rows[1:]:
        def cell(key):
            v = r[ix[key]]
            return str(v).strip() if v is not None else ""

        ticket, email = cell("ticket"), cell("email").lower()
        if not ticket or not email or ticket in seen:
            continue
        seen.add(ticket)
        name = " ".join(x for x in (cell("first"), cell("last")) if x)
        chapter = CHAPTER_CORRECTIONS.get(ticket, cell("chapter"))
        company = by_ticket.get(ticket, "")
        if company:
            carried[0] += 1
        else:
            company = by_name.get(name_key(name), "")
            if company:
                carried[1] += 1
        out.append({
            "ticket": ticket,
            "name": name,
            "email": email,
            "phone": normalize_phone(cell("phone")),
            "company": company,
            "chapter": chapter,
            "classification": cell("classification"),
        })
    return out, carried


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("sheet")
    ap.add_argument("--sheet-name", default="Data Peserta",
                    help="worksheet holding the attendee list (default: Data Peserta)")
    args = ap.parse_args()

    people, carried = read_rows(pathlib.Path(args.sheet).expanduser(), args.sheet_name)
    if not people:
        sys.exit("no attendee rows found")

    # No hashing here any more. Every attendee starts on the committee's
    # single password, so the rows carry the seeder's placeholder and the API
    # turns it into a real hash — from SEED_PASSWORD — on the next start. One
    # bcrypt call at boot instead of nearly nine hundred at generation.
    for p in people:
        p["hash"] = PLACEHOLDER_HASH

    chapters = sorted({p["chapter"] for p in people if p["chapter"]})
    chapter_values = ",\n".join(f"    ({q(c)})" for c in chapters)
    people_values = ",\n".join(
        "    (" + ", ".join(q(p[k]) for k in
        ("ticket", "name", "email", "phone", "company", "chapter", "classification", "hash")) + ")"
        for p in people)
    ticket_values = ",\n".join(f"    ({q(p['ticket'])})" for p in people)
    no_company = sum(1 for p in people if not p["company"])

    sql = f"""-- The {len(people)} attendees of the ticketing export.
-- GENERATED by scripts/attendees_migration.py — edit the sheet, not this file.
--
-- This replaces the earlier attendee migration, which was built from an
-- export that listed the 4 September Gold Club tickets as separate rows. Every
-- one of those belonged to somebody who already held a 3 September ticket, so
-- seeding both gave that person two accounts; the newer export carries the
-- Gold Club as a ticket TYPE instead, one row per person per ticket.
--
-- One member per TICKET: the ticket number is the identity, and it is what
-- the attendee's QR carries. A buyer holding several tickets becomes several
-- attendees, because a booth owner buying four crew passes is four people
-- through the door and the sheet never collected the crew's names.
--
-- Every attendee starts on the committee's single password (SEED_PASSWORD).
-- The rows below carry a placeholder hash nobody can sign in with; the API's
-- seeder rewrites it on the next start, so the password lives in the
-- environment rather than in this file. must_set_password stays true, so the
-- app still makes each attendee choose their own on first sign-in.
--
-- {len(people) - no_company} of these carry a company, brought across by ticket number from the
-- export that had a Company Name column; the newer one dropped it. The
-- remaining {no_company} have none on file. Business Classification is not an
-- employer, so nothing is guessed here.

INSERT INTO chapters (name)
SELECT v.name FROM (VALUES
{chapter_values}
) AS v (name)
ON CONFLICT (name) DO NOTHING;

-- Idempotent: a ticket already in the database is left exactly as it is,
-- keeping its password, its booth stamps and its class registrations.
INSERT INTO users (name, email, password_hash, role, member_code, chapter, company,
                   phone, classification, must_set_password, ticket_number)
SELECT v.name, v.email, v.hash, 'member',
       'NATCON-2026-' || lpad(nextval('member_code_seq')::text, 5, '0'),
       v.chapter, v.company, v.phone, v.classification, true, v.ticket
FROM (VALUES
{people_values}
) AS v (ticket, name, email, phone, company, chapter, classification, hash)
WHERE NOT EXISTS (SELECT 1 FROM users u WHERE u.ticket_number = v.ticket);

-- A ticket that has left the sheet goes with it — that is how the duplicate
-- Gold Club accounts are cleared. Only when the attendee has nothing against
-- their name: one scan, one class seat or one check-in and the row stays, so
-- a database in use is never quietly emptied. Every table that references a
-- member cascades on delete, which is exactly why each is checked here.
DELETE FROM users u
WHERE u.role = 'member'
  AND u.ticket_number <> ''
  AND NOT EXISTS (
      SELECT 1 FROM (VALUES
{ticket_values}
      ) AS v (ticket) WHERE v.ticket = u.ticket_number)
  AND NOT EXISTS (SELECT 1 FROM visits x WHERE x.member_id = u.id)
  AND NOT EXISTS (SELECT 1 FROM seminar_registrations x WHERE x.member_id = u.id)
  AND NOT EXISTS (SELECT 1 FROM seminar_attendance x WHERE x.member_id = u.id)
  AND NOT EXISTS (SELECT 1 FROM networking_checkins x WHERE x.member_id = u.id)
  AND NOT EXISTS (SELECT 1 FROM draw_winners x WHERE x.member_id = u.id);
"""
    OUT.write_text(sql, encoding="utf-8")
    size = OUT.stat().st_size // 1024
    print(f"{OUT.relative_to(ROOT)} — {len(people)} attendees, {len(chapters)} chapters, {size} KB")
    print(f"  company: {carried[0]} carried by ticket, {carried[1]} by name, {no_company} left empty")


if __name__ == "__main__":
    main()
